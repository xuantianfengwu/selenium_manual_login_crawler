from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.chrome.service import Service as ChromeService
# from selenium.webdriver.edge.service import Service as EdgeService
# from selenium.webdriver.chrome.options import Options as ChromeOptions
# from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import os
import sys
import time
import logging
import pandas as pd
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

# 导入tkinter用于文件选择对话框
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

# 获取资源文件路径的辅助函数
def get_resource_path(relative_path):
    """获取资源文件的绝对路径，兼容开发环境和打包后的环境"""
    if hasattr(sys, '_MEIPASS'):
        # 打包后的环境
        return os.path.join(sys._MEIPASS, relative_path)
    else:
        # 开发环境
        return os.path.join(os.path.abspath("."), relative_path)

# 配置日志
log_file_path = "static/complete_example/crawler.log"
# 获取日志文件所在目录
log_dir = os.path.dirname(log_file_path)
# 如果目录不存在，则创建
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class AdvancedCrawler:
    def __init__(self, cookie_file_path="static/complete_example/cookies.json", headless=False):
        """初始化爬虫
        Args:
            cookie_file_path: Cookies保存路径
            headless: 是否使用无头模式
        """
        self.cookie_file_path = cookie_file_path
        self.headless = headless
        self.driver = None
        self.setup_browser_options()
    
    def setup_browser_options(self):
        """设置浏览器选项"""
        self.options = Options()
        # 基本配置
        self.options.add_argument("--disable-notifications")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument("--start-maximized")
        # 禁用图片加载，提升速度
        # self.options.add_argument("--blink-settings=imagesEnabled=false")
        # 设置用户代理
        self.options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
        # 无头模式
        if self.headless:
            self.options.add_argument("--headless")
            self.options.add_argument("--window-size=1920,1080")
            self.options.add_argument("--disable-gpu")
    
    def start_browser(self, chrome_driver_path=None):
        """启动浏览器        
        Args: chrome_driver_path: 本地ChromeDriver路径（可选），如果提供则使用本地驱动，否则尝试自动下载
        Returns: bool: 浏览器启动是否成功
        """
        try:
            logger.info("正在启动浏览器...")
            
            # 1. 优先使用提供的ChromeDriver路径
            if chrome_driver_path:
                if os.path.exists(chrome_driver_path):
                    logger.info(f"使用打包的ChromeDriver: {chrome_driver_path}")
                    service = Service(chrome_driver_path)                
            else:
                # 3. 最后尝试自动下载ChromeDriver
                logger.info("尝试自动下载ChromeDriver...")
                try:
                    from webdriver_manager.chrome import ChromeDriverManager
                    driver_path = ChromeDriverManager().install()
                    # from webdriver_manager.microsoft import EdgeDriverManager
                    # driver_path = EdgeDriverManager().install()
                    logger.info(f"driver_path:{driver_path}")
                    service = Service(driver_path)
                except Exception as e:
                    logger.error(f"自动下载ChromeDriver失败: {str(e)}")
                    raise
            self.driver = webdriver.Chrome(
                service=service,
                options=self.options
            )
            logger.info("浏览器启动成功")
            return True
        except Exception as e:
            logger.error(f"浏览器启动失败: {str(e)}")
            logger.info("请确保ChromeDriver可用或尝试手动指定路径")
            return False
    
    def navigate_to(self, url, wait_for_element=None, timeout=10):
        """导航到指定URL并可选择等待特定元素加载
        Args:
            url: 目标URL
            wait_for_element: 等待加载的元素选择器
            timeout: 等待超时时间（秒）
        Returns:
            bool: 是否导航成功
        """
        try:
            logger.info(f"正在访问: {url}")
            self.driver.get(url)
            # 如果指定了等待元素，则等待元素加载
            if wait_for_element:
                WebDriverWait(self.driver, timeout).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, wait_for_element))
                )
                logger.info(f"元素 {wait_for_element} 已加载")
            else:
                # 否则简单等待几秒
                time.sleep(3)
            return True
        except Exception as e:
            logger.error(f"导航失败: {str(e)}")
            return False
    
    def save_cookies(self):
        """保存当前会话的Cookies
        Returns:
            bool: 是否保存成功
        """
        try:
            if not self.driver:
                logger.error("无法保存Cookies，浏览器未启动")
                return False
            cookies = self.driver.get_cookies()
            # 确保目录存在
            cookie_dir = os.path.dirname(self.cookie_file_path)
            if cookie_dir and not os.path.exists(cookie_dir):
                os.makedirs(cookie_dir)
            with open(self.cookie_file_path, 'w', encoding='utf-8') as f:
                json.dump(cookies, f, ensure_ascii=False, indent=4)
            logger.info(f"成功保存 {len(cookies)} 个Cookies到 {os.path.abspath(self.cookie_file_path)}")
            return True
        except Exception as e:
            logger.error(f"保存Cookies失败: {str(e)}")
            return False
    
    def load_cookies(self, url=None):
        """加载Cookies到当前会话
        Args: url: 如果提供，会先访问该URL再加载Cookies
        Returns: bool: 是否加载成功
        """
        try:
            if not self.driver:
                logger.error("无法加载Cookies，浏览器未启动")
                return False
            # 如果提供了URL，先访问
            if url:
                self.navigate_to(url)            
            # 检查文件是否存在
            if not os.path.exists(self.cookie_file_path):
                logger.error(f"Cookies文件不存在: {os.path.abspath(self.cookie_file_path)}")
                return False
            # 读取Cookies
            with open(self.cookie_file_path, 'r', encoding='utf-8') as f:
                cookies = json.load(f)
            # 清除现有Cookies
            self.driver.delete_all_cookies()
            # 添加Cookies
            success_count = 0
            for cookie in cookies:
                try:
                    # 处理可能的缺失字段
                    if 'expiry' in cookie and isinstance(cookie['expiry'], float):
                        cookie['expiry'] = int(cookie['expiry'])
                    self.driver.add_cookie(cookie)
                    success_count += 1
                except Exception as e:
                    logger.warning(f"添加Cookie失败: {str(e)}，跳过该Cookie")            
            # 刷新页面以应用Cookies
            self.driver.refresh()
            logger.info(f"成功加载 {success_count}/{len(cookies)} 个Cookies")
            return True
        except Exception as e:
            logger.error(f"加载Cookies失败: {str(e)}")
            return False
    
    def take_screenshot(self, filename="screenshot.png"):
        """截取当前页面的屏幕截图        
        Args: filename: 保存的文件名            
        Returns: bool: 是否截图成功
        """
        try:
            if not self.driver:
                logger.error("无法截图，浏览器未启动")
                return False
            self.driver.save_screenshot(filename)
            logger.info(f"截图已保存到 {os.path.abspath(filename)}")
            return True
        except Exception as e:
            logger.error(f"截图失败: {str(e)}")
            return False
    
    def close_browser(self):
        """关闭浏览器"""
        try:
            if self.driver:
                logger.info("正在关闭浏览器...")
                self.driver.quit()
                self.driver = None
                logger.info("浏览器已关闭")
        except Exception as e:
            logger.error(f"关闭浏览器失败: {str(e)}")

def demo_workflow():
    """演示完整的爬虫工作流程"""
    # --------------------------
    # 1. 文件选择阶段
    # --------------------------
    logger.info("===== 开始演示完整爬虫流程 =====")
    logger.info("\n----- 文件选择阶段 -----")
    # 创建tkinter窗口（隐藏）
    root = tk.Tk()
    root.withdraw()
    # 弹出文件选择对话框
    input_file_path = filedialog.askopenfilename(
        title="选择公司列表文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    # 检查用户是否选择了文件
    if not input_file_path:
        logger.error("用户未选择输入文件，程序终止")
        messagebox.showerror("错误", "未选择输入文件，程序终止")
        return    
    # 验证文件是否存在且为.xlsx格式
    if not os.path.exists(input_file_path):
        logger.error(f"所选文件不存在: {input_file_path}")
        messagebox.showerror("错误", f"所选文件不存在: {input_file_path}")
        return    
    if not input_file_path.endswith('.xlsx'):
        logger.error(f"所选文件不是有效的Excel文件(.xlsx): {input_file_path}")
        messagebox.showerror("错误", f"所选文件不是有效的Excel文件(.xlsx): {input_file_path}")
        return   
    # 1. 读取公司列表 
    logger.info(f"已选择输入文件: {input_file_path}")
    print(f"已选择输入文件: {input_file_path}")
    df = pd.read_excel(input_file_path)
    company_names = df['工商全称'].dropna().tolist()
    print(f'公司列表共 {len(company_names)} 家')

    # 网站URL示例（使用百度作为演示）
    target_url = "https://xunkebao.baidu.com"

    # Step1：首次访问并保存Cookies
    logger.info("\n----- 第一阶段：首次访问并保存Cookies -----")
    crawler = AdvancedCrawler()
    # chrome_driver_path = './static/chromedriver-win64/chromedriver.exe'
    # if crawler.start_browser(chrome_driver_path=chrome_driver_path):
    crawler.start_browser()
    crawler.navigate_to(target_url) # 1-1: 访问网站
    input("请在浏览器中执行任何需要的操作（如登录），完成后按Enter键继续...") # 1-2: 提示用户可以手动登录（如果需要）
    crawler.take_screenshot("static/complete_example/before_save_cookies.png") # 1-3: 截图记录当前状态
    crawler.save_cookies() # 1-4: 保存Cookies
    # crawler.close_browser() # 1-5: 关闭浏览器
    
    # Step2：使用保存的Cookies重新访问
    new_crawler = crawler
    # logger.info("\n----- 第二阶段：使用保存的Cookies重新访问 -----")
    # new_crawler = AdvancedCrawler()
    # try:
    #     if new_crawler.start_browser(chrome_driver_path=chrome_driver_path):
    #         # 1. 访问网站，然后加载Cookies
    #         new_crawler.navigate_to(target_url)
    #         new_crawler.load_cookies()
    #         # 2. 等待页面更新
    #         time.sleep(3)
    #         # 3. 截图验证状态
    #         new_crawler.take_screenshot("static/complete_example/after_load_cookies.png")
    #         logger.info("Cookies已加载，您可以在浏览器中验证状态是否保持（如登录状态）")
    #         input("验证完成后按Enter键继续...")
    # except Exception as e:
    #     logger.error(f"加载Cookies失败: {str(e)}")
    
    # Step3: 顺序爬取公司信息
    # try:
    # 2. 创建结果文件（如存在则直接读取）
    # output_file_path = os.path.join(os.path.dirname(__file__),'static','complete_example','output.xlsx')
    output_file_path = 'static/complete_example/output.xlsx'
    print(f'产出文件路径: {output_file_path}')
    if not os.path.exists(output_file_path):
        logger.info('未发现产出文件，将新建')
        wb = Workbook()
        ws_phone = wb.active
        ws_phone.title = "手机"
        ws_tel = wb.create_sheet("座机") # 创建第二个sheet "邮箱"
        ws_email = wb.create_sheet("邮箱") # 创建第二个sheet "邮箱"
        ws_qq = wb.create_sheet("QQ") # 创建第三个sheet "QQ"
        wb.save(output_file_path) # 保存工作簿
        existed_company_names = []
    else:
        logger.info('已发现产出文件，将直接读取')
        phone_df = pd.read_excel(output_file_path, sheet_name='手机', header=None)
        if len(phone_df)==0:
            existed_company_names = []
        else:
            existed_company_names = set([v.split('|')[0] for v in phone_df[0]])
        logger.info(f'已存在 {len(existed_company_names)} 家公司')

    # 3. 顺序遍历爬取
    for i, company_name in enumerate(company_names):
        if i<815:
            continue
        if company_name in existed_company_names:
            logger.info(f'公司 {company_name} 已存在，跳过')
            continue
        logger.info(f'开始处理公司 {i+1}/{len(company_names)}: {company_name}')
        # 3-1. 搜索公司
        search_input = new_crawler.driver.find_element(By.CSS_SELECTOR, "div.search-input-wrap > section > div > div > div > div > input")
        search_btn = new_crawler.driver.find_element(By.CSS_SELECTOR, "div.search-input-wrap > section > div > button.el-button.el-button--primary.search-btn")

        search_input.clear()
        search_input.send_keys(company_name)
        search_btn.click()
        # 3-2. 等待搜索结果加载
        time.sleep(3)

        # 3-3. 截图记录当前状态, 点开公司详情页
        search_res_num = new_crawler.driver.find_element(By.CSS_SELECTOR,
                                                         'div.middle-bar > div.info > span:nth-child(1) > em')
        if search_res_num.text == '0':
            continue
        # new_crawler.take_screenshot(f"static/complete_example/{company_name.replace(' ','_')}.png")
        company_btns = new_crawler.driver.find_elements(By.CSS_SELECTOR, "h6.company-name")
        is_find_btn = 0
        for btn in company_btns:
            dealed_company_name = company_name.replace('（','(').replace('）',')')
            if btn.text.strip() == company_name or btn.text.strip() == dealed_company_name:
                is_find_btn = 1
                btn.click()
                break
        if is_find_btn == 0:
            continue
        time.sleep(2)

        # 3-3 点击公司详情页的【立即查看】按钮
        check_btns = new_crawler.driver.find_elements(By.CSS_SELECTOR, "div.check > button")
        for btn in check_btns:
            if btn.text.strip() == "立即查看":
                # 滑动到对应按钮可见
                new_crawler.driver.execute_script("arguments[0].scrollIntoView(false);", btn)
                time.sleep(1)
                # 确保按钮可点击，再点击
                WebDriverWait(new_crawler.driver, 2).until(EC.element_to_be_clickable(btn))
                btn.click()
            break
        time.sleep(2)
        # 确定点击生效
        while True:
            time.sleep(2)
            check_btns = new_crawler.driver.find_elements(By.CSS_SELECTOR, "div.check > button")
            if len(check_btns)==0: # 找不到按钮，有时解锁后也会出现这个情况
                contact_div = new_crawler.driver.find_element(By.CSS_SELECTOR, "div.contact-item > div > div.p")
                contact_text = contact_div.text.strip()
                if len(contact_text)>0 and '*' not in contact_text:
                    break
                continue
            elif check_btns[0].text == "空错号检测":
                break
            else:
                continue

        # 3-4. 保存网页源代码，解析公司联系信息
        page_source = new_crawler.driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")
        modules = soup.select("#app > div > div.el-overlay.popMenu-box > div > section > div.container > div.coding-box > div > div")
        result_dict = {}
        for m in modules:
            title_text = m.select_one('div.list-item-title').get_text()
            title_class = title_text.split(' ')[0]
            item_text_lst = [t.get_text(separator='|') for t in m.select('div.contact-item')]
            result_dict[title_class] = [f'{company_name}|{title_text}|{i}' for i in item_text_lst]
            # print(f'{title_class}: {result_dict[title_class]}')
        # 3-5. 保存公司信息到结果文件
        wb = load_workbook(output_file_path)
        for n in ('手机', '座机', '邮箱', 'QQ'):
            ws = wb[n]
            for item in result_dict.get(n,[]):
                if isinstance(item, str):
                    ws.append([item])
        wb.save(output_file_path)
        time.sleep(2)
        # 3-6. 返回搜索页
        backhome_btn = new_crawler.driver.find_element(By.CSS_SELECTOR, "div.el-overlay.popMenu-box > div > section > div.pack-up")
        backhome_btn.click()
        time.sleep(1)
    # except Exception as e:
    #     logger.error(f"顺序爬取公司信息失败: {str(e)}")
    #     time.sleep(10)

    logger.info("\n===== 爬虫流程演示完成 =====")

if __name__ == "__main__":
    # 运行演示流程
    demo_workflow()